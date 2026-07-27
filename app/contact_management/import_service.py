# app/contact_management/import_service.py

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Iterator, Optional

from app.contact_management import repositories
from app.contact_management.repositories import ContactValidationError


HEADER_ALIASES = {
    "hubspot_record_id": {
        "hubspot record id",
        "record id",
        "hs record id",
        "contact id",
        "hubspot id",
        "id",
    },
    "first_name": {
        "first name",
        "firstname",
        "first",
    },
    "last_name": {
        "last name",
        "lastname",
        "last",
    },
    "email": {
        "email",
        "email address",
        "e mail",
        "e mail address",
    },
    "company_name": {
        "associated company",
        "associated company name",
        "primary associated company",
        "primary associated company name",
        "company name",
        "company",
    },
}

REQUIRED_HUBSPOT_FIELDS = {
    "company_name",
    "first_name",
    "last_name",
    "email",
}

HEADER_SCAN_LIMIT = 25


@dataclass(frozen=True)
class HubspotContactRow:
    hubspot_record_id: Optional[str]
    company_name: str
    first_name: str
    last_name: str
    email: str


@dataclass
class ImportCounts:
    inserted: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    for character in ("_", "-", "/", "\\"):
        text = text.replace(character, " ")
    return " ".join(text.split())


def header_mapping(headers: Iterable[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}

    for index, header in enumerate(headers):
        normalized = normalize_header(header)

        for field_name, aliases in HEADER_ALIASES.items():
            if field_name not in mapping and normalized in aliases:
                mapping[field_name] = index

    return mapping


def _cell(
    row: list[Any],
    mapping: dict[str, int],
    field_name: str,
) -> str:
    index = mapping.get(field_name)

    if index is None or index >= len(row):
        return ""

    return str(row[index] or "").strip()


def normalize_hubspot_row(
    row: Iterable[Any],
    mapping: dict[str, int],
) -> Optional[HubspotContactRow]:
    values = list(row)

    if not any(str(value or "").strip() for value in values):
        return None

    return HubspotContactRow(
        hubspot_record_id=repositories.normalize_hubspot_record_id(
            _cell(values, mapping, "hubspot_record_id")
        ),
        company_name=repositories._clean_required(
            _cell(values, mapping, "company_name"),
            "Company name",
        ),
        first_name=repositories._clean_required(
            _cell(values, mapping, "first_name"),
            "First name",
        ),
        last_name=repositories._clean_required(
            _cell(values, mapping, "last_name"),
            "Last name",
        ),
        email=repositories.normalize_email(
            _cell(values, mapping, "email")
        ),
    )


def _preferred_worksheet(workbook: Any) -> Any:
    for worksheet in workbook.worksheets:
        if normalize_header(worksheet.title) == "all contacts":
            return worksheet

    return workbook.active


def _find_header_row(
    worksheet: Any,
) -> tuple[dict[str, int], Iterator[tuple[Any, ...]]]:
    rows = worksheet.iter_rows(values_only=True)

    for row_number, row in enumerate(rows, start=1):
        mapping = header_mapping(row)

        if REQUIRED_HUBSPOT_FIELDS.issubset(mapping):
            return mapping, rows

        if row_number >= HEADER_SCAN_LIMIT:
            break

    missing = ", ".join(sorted(REQUIRED_HUBSPOT_FIELDS))
    raise ContactValidationError(
        "Could not locate a valid HubSpot header row. "
        f"Expected columns equivalent to: {missing}."
    )


def _open_hubspot_rows(
    workbook: Any,
) -> tuple[dict[str, int], Iterator[tuple[Any, ...]]]:
    preferred = _preferred_worksheet(workbook)
    worksheets = [preferred]

    worksheets.extend(
        worksheet
        for worksheet in workbook.worksheets
        if worksheet is not preferred
    )

    errors: list[str] = []

    for worksheet in worksheets:
        try:
            return _find_header_row(worksheet)
        except ContactValidationError as exc:
            errors.append(f"{worksheet.title}: {exc}")

    raise ContactValidationError(
        "No worksheet contained the required HubSpot contact columns. "
        + " | ".join(errors)
    )


def iter_hubspot_excel_rows(
    path: str | Path,
) -> Iterable[HubspotContactRow]:
    from openpyxl import load_workbook

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    try:
        mapping, rows = _open_hubspot_rows(workbook)

        for row in rows:
            normalized = normalize_hubspot_row(row, mapping)

            if normalized is not None:
                yield normalized
    finally:
        workbook.close()


def import_hubspot_contacts(
    path: str | Path,
    *,
    conn: object | None = None,
) -> ImportCounts:
    from openpyxl import load_workbook

    counts = ImportCounts()
    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    try:
        mapping, rows = _open_hubspot_rows(workbook)

        for excel_row in rows:
            try:
                row = normalize_hubspot_row(excel_row, mapping)

                if row is None:
                    counts.skipped += 1
                    continue

                before = repositories.get_client_contact_by_hubspot_or_email(
                    hubspot_record_id=row.hubspot_record_id,
                    email=row.email,
                    conn=conn,
                )

                repositories.upsert_client_contact_from_hubspot(
                    hubspot_record_id=row.hubspot_record_id,
                    company_name=row.company_name,
                    first_name=row.first_name,
                    last_name=row.last_name,
                    email=row.email,
                    conn=conn,
                )

                if before:
                    counts.updated += 1
                else:
                    counts.inserted += 1

            except ContactValidationError:
                counts.skipped += 1
            except Exception:
                counts.errors += 1
    finally:
        workbook.close()

    return counts


def load_internal_seed_rows(
    path: str | Path,
) -> list[dict[str, Any]]:
    seed_path = Path(path)

    if seed_path.suffix.lower() == ".json":
        data = json.loads(seed_path.read_text(encoding="utf-8"))

        if not isinstance(data, list):
            raise ContactValidationError(
                "Internal contact JSON seed must be a list."
            )

        return [
            dict(item)
            for item in data
            if isinstance(item, dict)
        ]

    if seed_path.suffix.lower() == ".csv":
        with seed_path.open(
            newline="",
            encoding="utf-8-sig",
        ) as handle:
            return [
                dict(row)
                for row in csv.DictReader(handle)
            ]

    raise ContactValidationError(
        "Internal contact seed must be a JSON or CSV file."
    )


def seed_internal_contacts(
    path: str | Path,
    *,
    conn: object | None = None,
) -> ImportCounts:
    counts = ImportCounts()

    for row in load_internal_seed_rows(path):
        try:
            before = repositories.get_internal_contact_by_email(
                row.get("email"),
                conn=conn,
            )

            if before:
                repositories.update_internal_contact(
                    before.id,
                    name=row.get("name"),
                    title=row.get("title"),
                    email=row.get("email"),
                    active=_seed_active_value(
                        row.get("active", True)
                    ),
                    conn=conn,
                )
                counts.updated += 1
            else:
                repositories.create_internal_contact(
                    name=row.get("name"),
                    title=row.get("title"),
                    email=row.get("email"),
                    active=_seed_active_value(
                        row.get("active", True)
                    ),
                    conn=conn,
                )
                counts.inserted += 1

        except ContactValidationError:
            counts.skipped += 1
        except Exception:
            counts.errors += 1

    return counts


def _seed_active_value(value: Any) -> bool:
    if isinstance(value, bool):
        return value

    normalized = str(value).strip().lower()

    if normalized in {
        "false",
        "0",
        "no",
        "n",
        "inactive",
    }:
        return False

    return True